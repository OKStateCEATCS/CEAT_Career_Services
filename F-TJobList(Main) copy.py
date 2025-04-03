from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import date
import re
import warnings
import os
import shutil
import time



#declaring credentials for HireOSUGrads
user_name = "ceatcs@okstate.edu"
pass_word = "hireCEAT2022!"
file_name = "Full-Time Job List (" + date.today().strftime("%B, %Y") + ").xlsx"

#This function allows us to automatically organize the excel files by yearly folders
def format_folder():
    outer_dir="K:\\studentservices\\crsvc_sh\\Job List\\"
    inner_dir="Job List {}\\"
    inner_dir=inner_dir.format(date.today().strftime("%Y"))
    #this is combining our two path parts into one and setting that as our path
    path=os.path.join(outer_dir,inner_dir)
    global year_dir
    year_dir=path
    #we are either making our folder or it already exists and we go to grab_excel
    try:
        os.mkdir(path)
        grab_excel()
    except FileExistsError:
        grab_excel()
    except FileNotFoundError:
        print("The given path is invalid, check the outer directory for changes")

#this function utilizes selenium to access HireOSUGrads and download the job list
def grab_excel():
    driver = webdriver.Chrome()
    
    #searching for the url
    driver.get("https://okstate.admin.12twenty.com/CustomReports#/customReports/163898")
    
    time.sleep(1)
    
    #finding the username input via html and inputting the credentials
    WebDriverWait(driver,20)
    user_input = driver.find_element(By.NAME, 'Username')
    user_input.send_keys(user_name)
    
    #finding the password input via html and inputting the credentials and logging in
    WebDriverWait(driver,20)
    pass_input = driver.find_element(By.NAME, 'Password')
    pass_input.send_keys(pass_word, Keys.ENTER)
    
    #fullscrrens the window to find elements
    driver.maximize_window()

    #Finding and waiting for the action button to be clickable and then clicking it
    btn_element=WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH, '//button[contains(text(),"Action")]')))
    btn_element.click()

    #Finding and waiting for the export button to be clickable and then clicking it
    link_element=WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH, '//a[contains(text(),"Export")]')))
    link_element.click()
    
    #give time for function to execute
    time.sleep(2)
    
#function that grabs CEAT Job Report from downloads and moves it to job list folder with new name
def move_excel():
    try:
        shutil.move(r'C:\\Users\\ceatcs\\Downloads\\CEAT Full-Time Job Report.xlsx',r''+year_dir+file_name)
    except:
        print('Error finding file... Trying again.')
        move_excel()
    
#function that organizes Job List file
def edit_excel():
    arch_majors = ['Bach - Architecture', 'Bach - Architectural Engineering - Construction Project Management', 'Bach - Architectural Engineering - Structures', 'Bach - Architectural Engineering']
    bae_majors = ['Bach - Biosystems Engineering', 'Bach - Biosystems Engineering - Bioprocessing and Food Processing', 'Bach - Biosystems Engineering - Environmental and Natural Resources', 'Bach - Biosystems Engineering - Machine Systems and Agricultural Engineering', 'Bach - Biosystems Engineering - Pre-Medical', 'Mast - Biosystems Engineering', 'Mast - Environmental Engineering', 'Doc - Biosystems Engineering']   
    chem_majors = ['Bach - Chemical Engineering', 'Bach - Chemical Engineering - Biomedical/Biochemical', 'Bach - Chemical Engineering - Pre-Medical', 'Mast - Chemical Engineering', 'Mast - Materials Science and Engineering', 'Doc - Chemical Engineering','Doc - Materials Science and Engineering' ]
    cive_majors = ['Bach - Civil Engineering', 'Bach - Civil Engineering - Environmental', 'Mast - Civil Engineering', 'Doc - Civil Engineering']
    ecen_majors = ['Bach - Computer Engineering', 'Bach - Computer Engineering - Software Engineering', 'Bach - Electrical Engineering', 'Mast - Electrical Engineering', 'Mast - Electrical Engineering - Control Systems', 'Mast - Electrical Engineering - Optics and Photonics', 'Doc - Electrical Engineering']
    iem_majors = ['Bach - Industrial Engineering and Management', 'Mast - Industrial Engineering and Management', 'Mast - Engineering and Technology Management', 'Mast - Industrial Engineering and Management - Operations Research and Analytics', 'Mast - Industrial Engineering and Management - Supply Chain and Logistics', 'Doc - Industrial Engineering and Management']
    mae_majors = ['Bach - Aerospace Engineering','Bach - Mechanical Engineering', 'Bach - Mechanical Engineering - Pre-Medical', 'Bach - Mechanical Engineering - Fire Protection Systems', 'Bach - Mechanical Engineering - Petroleum', 'Mast - Mechanical and Aerospace Engineering', 'Mast - Mechanical and Aerospace Engineering - Unmanned Aerial Systems', 'Mast - Materials Science and Engineering', 'Mast - Petroleum Engineering', 'Doc - Materials Science and Engineering', 'Doc - Mechanical and Aerospace Engineering', 'Doc - Mechanical and Aerospace Engineering - Unmanned Aerial Systems', 'Doc - Petroleum Engineering']
    tech_majors = ['Bach - Construction Engineering Technology', 'Bach - Construction Engineering Technology - Building', 'Bach - Construction Engineering Technology - Heavy', 'Bach - Electrical Engineering Technology', 'Bach - Electrical Engineering Technology - Computer', 'Bach - Mechanical Engineering Technology', 'Bach - Mechatronics and Robotics', 'Mast - Engineering Technology - Fire Safety and Explosion Protection', 'Mast - Engineering Technology', 'Mast - Engineering Technology - Mechatronics and Robotics']
    #this ignores a warning regarding the look of the sheet
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module=re.escape('openpyxl.styles.stylesheet'))
        book = load_workbook(r''+year_dir+file_name) 
    ws=book['CEAT Full-Time Job Report']
    #creating excel sheets at bottom of page
    
    sheet1 = book.create_sheet('ARCH')
    sheet1.title = 'ARCH'
    sheet2 = book.create_sheet('BAE')
    sheet2.title = 'BAE'
    sheet3 = book.create_sheet('CHEM')
    sheet3.title = 'CHEM'
    sheet4 = book.create_sheet('CIVE')
    sheet4.title = 'CIVE'
    sheet5 = book.create_sheet('ECEN')
    sheet5.title = 'ECEN'
    sheet6 = book.create_sheet('IEM')
    sheet6.title = 'IEM'
    sheet7 = book.create_sheet('MAE')
    sheet7.title = 'MAE'
    sheet8 = book.create_sheet('TECHNOLOGY')
    sheet8.title = 'TECHNOLOGY'
    
    #assigning width to each column
    column_list = 'ABCDEFG'
    for i in column_list:
        ws.column_dimensions[i].width = 40
        sheet1.column_dimensions[i].width = 40
        sheet2.column_dimensions[i].width = 40
        sheet3.column_dimensions[i].width = 40
        sheet4.column_dimensions[i].width = 40
        sheet5.column_dimensions[i].width = 40
        sheet6.column_dimensions[i].width = 40
        sheet7.column_dimensions[i].width = 40
        sheet8.column_dimensions[i].width = 40
    
    #giving each sheet same heading in row 1 (i.e. Job Title, Company Name, etc.)
    for i in range(1,8):
        sheet1.cell(row=1, column=i).value = ws.cell(row=1,column=i).value
        sheet2.cell(row=1, column=i).value = ws.cell(row=1,column=i).value
        sheet3.cell(row=1, column=i).value = ws.cell(row=1,column=i).value
        sheet4.cell(row=1, column=i).value = ws.cell(row=1,column=i).value
        sheet5.cell(row=1, column=i).value = ws.cell(row=1,column=i).value
        sheet6.cell(row=1, column=i).value = ws.cell(row=1,column=i).value
        sheet7.cell(row=1, column=i).value = ws.cell(row=1,column=i).value
        sheet8.cell(row=1, column=i).value = ws.cell(row=1,column=i).value

    #function that puts each job listing in designated sheet. 
    def sheet_organize(sheet, diff_majors):
        current_row = 0 
        next_row = 1
        for cell in ws['D']:
            current_row = current_row + 1
            temp_val = cell.value
            if temp_val is not None:
                temp_arr = temp_val.split(', ')
            count_val = 0
            for i in range(len(diff_majors)):
                if diff_majors[i] in temp_arr:
                    count_val = count_val + 1
            if count_val > 0:
                next_row = next_row + 1
                sheet.cell(row=next_row, column=1).value = ws.cell(row=current_row, column=1).value
                sheet.cell(row=next_row, column=2).value = ws.cell(row=current_row, column=2).value
                sheet.cell(row=next_row, column=3).value = ws.cell(row=current_row, column=3).value
                sheet.cell(row=next_row, column=4).value = ws.cell(row=current_row, column=4).value
                sheet.cell(row=next_row, column=5).value = ws.cell(row=current_row, column=5).value
                sheet.cell(row=next_row, column=6).value = ws.cell(row=current_row, column=6).value
                sheet.cell(row=next_row, column=7).value = ws.cell(row=current_row, column=7).value
        
    sheet_organize(sheet1, arch_majors)
    sheet_organize(sheet2, bae_majors)
    sheet_organize(sheet3, chem_majors)
    sheet_organize(sheet4, cive_majors)
    sheet_organize(sheet5, ecen_majors)
    sheet_organize(sheet6, iem_majors)
    sheet_organize(sheet7, mae_majors)
    sheet_organize(sheet8, tech_majors)
    
    book.save(r''+year_dir+file_name)   
    
    
format_folder()
time.sleep(3)
move_excel()
print("FOUND THE FILE.")
edit_excel()
time.sleep(1)