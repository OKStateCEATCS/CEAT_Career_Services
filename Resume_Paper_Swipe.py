from datetime import date, timedelta
import openpyxl

#This loads the sheet for viewing/editing
ResumeScan = openpyxl.load_workbook('J:\\studentservices\\crsvc_sh\\Resume Paper Scan\\Resume Paper Scan (DONT OPEN).xlsx')

#This allows us to read the sheet
sheet1 = ResumeScan.active


#This searches for the inputted ID number in Column A
def search():
    #This gets us the students card ID to cross reference
    ID = input("Swipe ID Please or type 'exit' to close: ")
    global ID_itr
    ID_itr = ''
    if ID.lower() == "exit":
        exit()
    else:
        for char in ID:
            if char.isdigit():
                ID_itr +=char
    #This is the parameter to give paper
    global GivePaper
    GivePaper=True
    for col in sheet1.iter_cols(max_col=1):
        for cell in col:
            #This compares the value in a given cell to the ID
            if (cell.value == ID_itr):
                ID_row = cell.row
                #if they match we will search through column B for a date within the last week
                for col in sheet1.iter_cols(min_col=2, min_row=ID_row, max_row=ID_row, max_col=2):
                    for cell in col:
                        #for each value in column B we will check if it matches any of the dates within the last week up until today
                        count=7
                        while (count >= 0):
                            PastDate = date.today() - timedelta(days=count)
                            #this formats our date into the month/day/year
                            CheckDate = PastDate.strftime("%m/%d/%Y")
                            #if the date matches
                            if (cell.value == CheckDate):
                                GivePaper = False
                            count = count-1
        log_id()

def log_id():
    if (GivePaper==True):
        sheet1.cell(row=sheet1.max_row +1, column=1, value=ID_itr)
        sheet1.cell(row=sheet1.max_row, column=2, value=date.today().strftime("%m/%d/%Y"))
        print("Information added, please collect your paper.")
    else:
        print("Paper has been collected recently, come back in a few days and try again!")
                
    ResumeScan.save('J:\\studentservices\\crsvc_sh\\Resume Paper Scan\\Resume Paper Scan (DONT OPEN).xlsx')
    search()

search()