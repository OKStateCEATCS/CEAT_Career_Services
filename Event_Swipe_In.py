from datetime import date
import openpyxl


def file_creation():
    query_name = "CEAT Student Data ({}).xlsx"
    query_path= f"{date.today().strftime("%Y")} Queries"
    if 1<=int(date.today().strftime("%m"))<=5:
        query_name = query_name.format("Spring, "+date.today().strftime("%Y"))
    elif 6<=int(date.today().strftime("%m"))<=7:
        query_name = query_name.format("Summer, "+date.today().strftime("%Y"))
    else:
        query_name = query_name.format("Fall, "+date.today().strftime("%Y"))
    #This loads the sheet for viewing/editing
    StudentData = openpyxl.load_workbook(f"J:\\studentservices\\crsvc_sh\\Queries\\{query_path}\\{query_name}")
    global query_sheet
    query_sheet = StudentData.active
    
    global event_name
    event_name = input("Please input the event name: ")
    global EventList
    try:
        EventList= openpyxl.load_workbook(f"J:\\studentservices\\crsvc_sh\\Event Check In\\{event_name}.xlsx")
    except FileNotFoundError:
        EventList = openpyxl.Workbook()
    global event_sheet
    event_sheet = EventList.active
    event_sheet.title = "Event Attendance"
    get_cwid()


#This iterates through the columns to find the iso id column
def get_cwid():
    #This gets us the students card ID to cross reference
    ID = input("Swipe ID Please or Type 'exit' To Close: ")
    ID_itr = ''
    if ID.lower() == "exit":
        exit()
    else:
        for char in ID:
            if char.isdigit():
                ID_itr +=char
    #input("Press enter to continue")
    for col in query_sheet.iter_cols():
        column_name = col[0].value
        if "Card ID" in column_name:
            iso_col=col
        if "Banner ID" in column_name:
            cwid_col=col
            #This searches through the rows in the iso column for the students iso number
    for cell in iso_col:
        if cell.value == ID_itr:
            ID_row = cell.row
    global cwid
    try:
        for cell in cwid_col: 
            if cell.row == ID_row:
                cwid_col=cell.column
                cwid=query_sheet.cell(row=ID_row, column=cwid_col).value
                log_cwid()
    except(UnboundLocalError):
        cwid=input("Please enter cwid(i.e. A12345678): ")
        log_cwid()

def log_cwid():
    event_sheet.cell(row=event_sheet.max_row +1, column=1, value = cwid)
    EventList.save(f"J:\\studentservices\\crsvc_sh\\Event Check In\\{event_name}.xlsx")
    print("ID Saved")
    get_cwid()
    
        
file_creation()