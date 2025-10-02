from openpyxl import load_workbook
from datetime import date
from tkinter import *
import tkinter as tk
import os
import re
import warnings


#declaring query file name
file_name = "CEAT Student Data ({}).xlsx"
if 1<=int(date.today().strftime("%m"))<=5:
    file_name = file_name.format("Spring, "+date.today().strftime("%Y"))
elif 6<=int(date.today().strftime("%m"))<=7:
    file_name = file_name.format("Summer, "+date.today().strftime("%Y"))
else:
    file_name = file_name.format("Fall, "+date.today().strftime("%Y"))




def search_info():
    def swipe_id():
        global FramesList
        global swipe_frame
        swipe_frame=tk.Frame(win, height=700, width=700)
        swipe_frame.grid(row=0, column=0)
        swipe_frame.rowconfigure(0, weight=1)
        swipe_frame.rowconfigure(1, weight=1)
        swipe_frame.columnconfigure(0, weight=1)
        FramesList.append(swipe_frame)
        
        id_l = tk.Label(swipe_frame, font=("Raleway", 20, "bold"), text="Please Swipe Your ID:")
        id_l.grid(row=0, column=0, sticky=(N,E,S,W))
        
        global iso_var    
        iso_var=tk.StringVar()
                
        id_swipe=tk.Entry(swipe_frame,textvariable=iso_var)
        id_swipe.focus_set()
        id_swipe.grid(row=1, column=0)  
        id_swipe.bind('<Return>', search_iso) 
        
        
    def search_iso(e):
        clear_frames()
        iso=iso_var.get()
        #implementing a text based kill switch to end the program
        if(iso=="exit"):
            exit()
        #Declaring book and ignoring a vanity related warning to clean up the output
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module=re.escape('openpyxl.styles.stylesheet'))
            book = load_workbook(r''+year_dir+file_name) 
        ws=book['Export']
        
        temp_iso=list([val for val in iso if val.isnumeric()])
        iso="".join(temp_iso)
                
        for column in ws.iter_cols():
            column_name = column[0].value
            if "Card ID" in column_name:
                iso_col=column
            if "Email" in column_name:
                email_col=column[0].col_idx
        
        
        current_row=0
        #If we find the iso, we confirm their info, otherwise we search using email
        trigger=0
        for cell in iso_col:
            current_row+=1
            temp_val = cell.value
            if (temp_val == iso):
                trigger=1
            #Retrieving Student's Email
                global email
                email_cell=ws.cell(row=current_row, column=email_col)
                email=email_cell.value
                info_confirmation()  
        if (trigger==0):
            search_email()
    
    
    def search_email():
        clear_frames()
        def assign():
            global email
            email=email_var.get()
            info_confirmation()
        
        global email_frame
        email_frame=tk.Frame(win, height=700, width=700)
        email_frame.grid(row=0, column=0)
        email_frame.rowconfigure(0, weight=2)
        email_frame.rowconfigure(1, weight=1)
        email_frame.rowconfigure(2, weight=2)
        email_frame.columnconfigure(0, weight=1)
        email_frame.columnconfigure(1, weight=1)
        FramesList.append(email_frame)
        
        global email_var
        email_var=tk.StringVar()

                
        email_pl=tk.Label(email_frame, font=("Raleway", 20, "bold"), text="Please input your email below:")
        email_pl.grid(row=0, column=0, columnspan=2, sticky=(S,E,W), padx=10, pady=10)
        
        el=tk.Label(email_frame, font=("Raleway", 16), text="Email:")
        el.grid(row=1, column=0, sticky=(N,E))
        
        e=tk.Entry(email_frame,textvariable=email_var)
        e.grid(row=1, column=1, sticky=(N,E,W))
            
        sub_btn=tk.Button(email_frame, font=("Raleway"), text="Submit", command=assign)
        sub_btn.grid(row=2, column=0, columnspan=2,sticky=(N,E,W))    
    
        
    def info_confirmation():
        clear_frames()
        
        global conf_frame
        conf_frame=tk.Frame(win, height=700, width=700)
        conf_frame.grid(row=0, column=0)
        conf_frame.rowconfigure(0, weight=1)
        conf_frame.rowconfigure(1, weight=1)
        conf_frame.columnconfigure(0, weight=1)
        conf_frame.columnconfigure(1, weight=1)
        FramesList.append(conf_frame)
        
        l = tk.Label(conf_frame, font=("Raleway", 20, "bold"), text="Is this your email?\n"+email)
        l.grid(row=0, column=0, columnspan=2)
        
        yes_btn=tk.Button(conf_frame, font=("Raleway"), text="Yes", command=send_email)
        yes_btn.grid(row=1, column=0, padx=10, pady=10)
        
        no_btn=tk.Button(conf_frame, font=("Raleway"), text="No", command=search_email)
        no_btn.grid(row=1, column=1, padx=10, pady=10)    
     
       
    #declaring window info 
    global win
    win=tk.Tk()
    win.title("Headshot Check-In")
    win.state("zoomed")
    win.rowconfigure(0, weight=1)
    win.columnconfigure(0, weight=1)
    
    global FramesList
    FramesList = []
    
    
    #declaring iso frame and related info
       
    swipe_id()
    win.mainloop()    

#Finding and Formatting the Query Folder
def format_folder():
    outer_dir="J:\\studentservices\\crsvc_sh\\Headshot Check In\\"
    inner_dir="{} Queries\\"
    inner_dir=inner_dir.format(date.today().strftime("%Y"))
    #this is combining our two path parts into one and setting that as our path
    path=os.path.join(outer_dir,inner_dir)
    global year_dir
    year_dir=path
    #we are either making our folder or it already exists and we go to grab_excel
    try:
        os.mkdir(path)
        search_info()
    except FileExistsError:
        search_info()
    except FileNotFoundError:
        print("The generated path is invalid; check the directories for changes.")
    
def clear_frames():
    global swipe_frame
    global email_frame
    global conf_frame
    for i in FramesList:
        i.destroy()
    
            
def restart():
    win.destroy()
    os.system('cls')
    search_info()
  
          
def send_email():
    clear_frames()
    #for Ethan
    
    #Put vv this vv at the end of your code
    restart()


format_folder()