from openpyxl import load_workbook
from datetime import date
from tkinter import *
import tkinter as tk
import os
import re
import warnings


#declaring query file name
file_name = "CEAT Student Data ({}).xlsx"
if 1<=int(date.today().strftime("%m"))<=12:
    file_name = file_name.format("Spring, "+date.today().strftime("%Y"))
#elif 6<=int(date.today().strftime("%m"))<=7:
#    file_name = file_name.format("Summer, "+date.today().strftime("%Y"))
#else:
#    file_name = file_name.format("Fall, "+date.today().strftime("%Y"))








def search_info(): 
    def search_iso(e):
        global iso
        print(iso)
        #implementing a text based kill switch to end the program
        if(iso=="exit"):
            exit()
        #Declaring book and ignoring a vanity related warning to clean up the output
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module=re.escape('openpyxl.styles.stylesheet'))
            book = load_workbook(r''+year_dir+file_name) 
        ws=book['Export']
        
        current_row=0

        #If we find the iso, we confirm their info, otherwise we search using email
        trigger=0
        for cell in ws['E']:
            current_row+=1
            temp_val = cell.value
            if (temp_val == iso):
                trigger=1
            #Retrieving Student's First Name
                global f_name
                fn_cell=ws.cell(row=current_row, column=2)
                f_name=fn_cell.value
            #Retrieving Student's Last Name       
                global l_name
                ln_cell=ws.cell(row=current_row, column=3)
                l_name=ln_cell.value  
            #Retrieving Student's Email
                global email
                email_cell=ws.cell(row=current_row, column=6)
                email=email_cell.value
                info_confirmation()  
        if (trigger==0):
            search_email()
    
    
    def search_email():
        global email_frame
        email_frame.grid(row=0, column=0)
        #Declaring book and ignoring a vanity related warning to clean up the output
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module=re.escape('openpyxl.styles.stylesheet'))
            book = load_workbook(r''+year_dir+file_name) 
        ws=book['Export']

        current_row=0

        #If we find it, we confirm their info, otherwise we get email
        for cell in ws['F']:
            current_row+=1
            temp_val = cell.value
            if temp_val == email:
            #Retrieving Student's First Name
                global f_name
                fn_cell=ws.cell(row=current_row, column=2)
                f_name=fn_cell.value
            #Retrieving Student's Last Name       
                global l_name
                ln_cell=ws.cell(row=current_row, column=3)
                l_name=ln_cell.value   
        info_confirmation()
    
        
    def info_confirmation():
        global win
        def confirmed():
            conf_frame.destroy()
            send_email()
        
        conf_frame=tk.Frame(win, bg='red')
        conf_frame.rowconfigure(0, weight=2)
        conf_frame.rowconfigure(3, weight=2)
        conf_frame.columnconfigure(0, weight=2)
        conf_frame.columnconfigure(3, weight=2)
        l = tk.Label(conf_frame, text="Is this your information?\n"+f_name+" "+l_name+"\n"+email)
        l.grid(row=1, column=0, columnspan=4)
        yes_btn=tk.Button(conf_frame, text="Yes", command=confirmed)
        yes_btn.grid(row=2, column=1)
        no_btn=tk.Button(conf_frame, text="No", command=search_email)
        no_btn.grid(row=2, column=2)    
     
       
    #declaring window info 
    global win
    win=tk.Tk()
    win.title("Headshot Check-In")
    win.state("zoomed")
    win.rowconfigure(0, weight=1)
    win.columnconfigure(0, weight=1)
    
    
    #declaring iso frame and related info
    iso_frame=tk.Frame(win, bg='cyan', height=700, width=700)
    iso_frame.grid(row=0, column=0)
    iso_frame.rowconfigure(0, weight=1)
    iso_frame.rowconfigure(1, weight=1)
    iso_frame.columnconfigure(0, weight=1)
    
    id_l = tk.Label(iso_frame, text="Please Swipe Your ID:")
    id_l.grid(row=0, column=0, sticky=(N,E,S,W))
        
    iso_var=tk.StringVar()
    global iso
    iso=iso_var.get()
    
    iso_entry=tk.Entry(iso_frame,textvariable=iso_var)
    iso_entry.focus_set()
    iso_entry.grid(row=1, column=0)  
    win.bind('<Return>', search_iso)
    
    
    #declaring email frame and related info
    global email_frame
    email_frame=tk.Frame(win, bg='green', height=700, width=700)

    email_frame.rowconfigure(0, weight=2)
    email_frame.rowconfigure(1, weight=1)
    email_frame.rowconfigure(2, weight=2)
    email_frame.columnconfigure(0, weight=1)
    email_frame.columnconfigure(1, weight=1)
    
    email_var=tk.StringVar()
    global email
    email=email_var.get()
    
    email_pl=tk.Label(email_frame,text="Please input your email below:")
    email_pl.grid(row=0, column=0, columnspan=1, sticky=(S,E,W), padx=10, pady=10)
    
    el=tk.Label(email_frame,text="Email:")
    el.grid(row=1, column=0, sticky=(N,E), padx=10, pady=10)
    
    e=tk.Entry(email_frame,textvariable=email_var)
    e.grid(row=1, column=1, sticky=(N,W), padx=10, pady=10)
        
    sub_btn=tk.Button(email_frame, text="Submit", command=search_email)
    sub_btn.grid(row=2, column=0, columnspan=2,sticky=(N,E,W), padx=10, pady=10 )
    
    
    win.mainloop()    

#Finding and Formatting the Query Folder
def format_folder():
    outer_dir="K:\\studentservices\\crsvc_sh\\Headshot Check In\\"
    inner_dir="{} Queries\\"
    inner_dir=inner_dir.format(date.today().strftime("%Y"))
    #this is combining our two path parts into one and setting that as our path
    path=os.path.join(outer_dir,inner_dir)
    global year_dir
    year_dir=path
    #we are either making our folder or it already exists and we go to grab_excel
    try:
        os.mkdir(path)
        search_info()
    except FileExistsError:
        search_info()
    except FileNotFoundError:
        print("The generated path is invalid; check the directories for changes.")
    
    
def restart():
    os.system('cls')
    search_info()
  
          
def send_email():
    #for Ethan
    
    #Put vv this vv at the end of your code
    restart()


format_folder()