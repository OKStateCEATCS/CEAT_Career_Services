from openpyxl import load_workbook
from datetime import datetime, date, timedelta

import openpyxl
import datetime

#This loads the sheet for viewing/editing
ResumeScan = openpyxl.load_workbook(r"J:\studentservices\crsvc_sh\Resume Paper Scan\Resume Paper Scan (DONT OPEN).xlsx" )

#This allows us to read the sheet
sheet1 = ResumeScan.active

#This gets us the students card ID to cross reference
ID = input("Swipe ID Please: ")

#This is the parameter to give paper
GivePaper=True

#This searches for the inputted ID number in Column A
for col in sheet1.iter_cols(max_col=1):
    for cell in col:
        #This compares the value in a given cell to the ID
        if (cell.value == ID):
            ID_row = cell.row
            #if they match we will search through column B for a date within the last week
            for col in sheet1.iter_cols(min_col=2, min_row=ID_row, max_row=ID_row, max_col=2):
                for cell in col:
                    #for each value in column B we will check if it matches any of the dates within the last week up until today
                    count=7
                    while (count >= 0):
                        PastDate = date.today() - timedelta(days=count)
                        #this formats our date into the month/day/year
                        CheckDate = PastDate.strftime("%m/%d/%Y")
                        #if the date matches
                        if (cell.value == CheckDate):
                            GivePaper = False
                        count = count-1
                    
if (GivePaper==True):
    idInfo = sheet1.cell(row=sheet1.max_row +1, column=1, value=ID)
    DateInfo = sheet1.cell(row=sheet1.max_row, column=2, value=date.today().strftime("%m/%d/%Y"))
    print("Information added, please collect your paper.")
else:
    print("Paper has been collected recently, come back in a few days and try again!")
            
ResumeScan.save(r"J:\studentservices\crsvc_sh\Resume Paper Scan\Resume Paper Scan (DONT OPEN).xlsx")

input("Press enter to exit")

        
