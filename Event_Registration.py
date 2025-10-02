from selenium import webdriver
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from datetime import date
from openpyxl import load_workbook
import re
import warnings
import os
import time

#Prompting the user for their username, password, and event link to log attendeance in slate
print("Greetings, you will need to input your SLATE Admin credentials as well as a link to the event"+"A passcode should be texted to you, please input it into the website.")
user_name = input("Please input your email: ")
pass_word = input("Please input your password: ")
event_link = input("Please input your event link: ")
#declaring the file name we want to find
file_name = "Student Data.xlsx"

#this loads the excel file, ignoring a cosmetic warning
with warnings.catch_warnings():
    warnings.filterwarnings("ignore", category=UserWarning, module=re.escape('openpyxl.styles.stylesheet'))
    book = load_workbook(r''+"J:\\studentservices\\crsvc_sh\\Event Check In\\"+file_name) 
#selecting the sheet we want to work in
ws=book['Sheet1']

#creating a huge list from the emails in the excel file
email_list= []
for cell in ws['A']:
    #this appends the cells email to the list
    email_list.append(cell.value)

#this function utilizes selenium to access HireOSUGrads and download the job list
def book_appt():
    #this declares our "driver", which is just the internet host we want. i.e chrome
    driver = webdriver.Chrome()
    
    #this action chain allows us to interact with the webpage later
    action = ActionChains(driver)
    
    #searching for the url
    driver.get(event_link)
    
    #waiting for the page to load fully before continuing
    time.sleep(1)
    
    #finding the username input via html and inputting the credentials
    WebDriverWait(driver,20)
    user_input = driver.find_element(By.NAME, 'username')
    user_input.send_keys(user_name)
    
    #finding the password input via html and inputting the credentials and logging in
    WebDriverWait(driver,20)
    pass_input = driver.find_element(By.NAME, 'password')
    pass_input.send_keys(pass_word, Keys.ENTER)
    
    #fullscrrens the window to find elements
    driver.maximize_window()

    #Finding and waiting for the passcode button to be clickable and then clicking it
    passcd_btn_element=WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH, '//button[contains(text(),"Send a passcode")]')))
    passcd_btn_element.click()
    
    #Finding and waiting for the export button to be clickable and then clicking it
    brw_btn_element=WebDriverWait(driver,1000).until(EC.element_to_be_clickable((By.ID, 'dont-trust-browser-button')))
    brw_btn_element.click()
    
 
    for email in email_list:
        #this sleep function is absolutely crucial. It allows for us to account for slow response time
        #if it gets removed or changed the program will crash because of the necessary components not being visible
        time.sleep(5)
        #waiting for the element to be able to be found
        rgstr_link_element=WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH, '//a[contains(text(),"New Registration")]')))
        #waiting for the element to be clickable and then clicking
        rgstr_link_element=WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH, '//a[contains(text(),"New Registration")]')))
        rgstr_link_element.click()
        
        #inputting the students email from the email lisyt created earlier    
        email_input = driver.find_element(By.ID, 'q_suggest')
        email_input.send_keys(email)
        
        time.sleep(1)
        #These functions allow us to click on the recommended student profile provided by slate based of the students email we input 
        action.click(on_element=email_input)
        action.send_keys(Keys.DOWN).perform()
        action.send_keys(Keys.ENTER).perform()
        
        try:
            sub_btn_element=driver.find_element(By.XPATH, '//button[contains(text(),"Submit")]')
            sub_btn_element.click()
            
        except StaleElementReferenceException:
            sub_btn_element=WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, '//button[contains(text(),"Submit")]')))
            sub_btn_element=WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH, '//button[contains(text(),"Submit")]')))
            sub_btn_element.click()
            
        except NoSuchElementException:
            upd_btn_element=WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, '//button[contains(text(),"Update")]')))
            upd_btn_element=WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH, '//button[contains(text(),"Update")]')))
            upd_btn_element.click()



book_appt()

